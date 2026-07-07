#!/usr/bin/env python3
"""Convertit le planning Excel en planning-data.json.
Prend automatiquement le fichier .xlsx le plus récemment ajouté (peu importe son nom),
sinon 'planning.xlsx'. Lancé par GitHub Actions à chaque mise à jour d'un fichier Excel.
"""
import json, glob, os, sys, subprocess
from datetime import datetime, timedelta
from openpyxl import load_workbook

def _all_xlsx():
    return [f for f in glob.glob("**/*.xlsx", recursive=True)
            if not os.path.basename(f).startswith("~$")]

def find_xlsx():
    files = _all_xlsx()
    if not files:
        print("Aucun fichier .xlsx trouvé dans le dépôt.", file=sys.stderr)
        sys.exit(1)
    if len(files) == 1:
        return files[0]
    # Plusieurs .xlsx : prendre le plus récemment modifié d'après l'historique git.
    try:
        order = subprocess.check_output(
            ["git", "log", "-80", "--name-only", "--pretty=format:", "--diff-filter=d"],
            text=True, stderr=subprocess.DEVNULL)
        seen = []
        for line in order.splitlines():
            line = line.strip()
            if line.endswith(".xlsx") and line in files and line not in seen:
                seen.append(line)
        if seen:
            return seen[0]
    except Exception:
        pass
    # Secours : 'planning.xlsx' si présent, sinon le plus récent par date de fichier.
    if os.path.exists("planning.xlsx"):
        return "planning.xlsx"
    return max(files, key=lambda f: os.path.getmtime(f))

def to_dt(v):
    if v is None:
        return None
    if isinstance(v, datetime):
        return v
    try:
        f = float(v)
        if 20000 <= f <= 80000:
            return datetime(1899, 12, 30) + timedelta(days=f)
    except (TypeError, ValueError):
        pass
    return None

def cell(v):
    return "" if v is None else str(v).strip()

def find_cols(rows):
    for i, r in enumerate(rows[:15]):
        idx = {}
        for j, c in enumerate(r):
            k = cell(c).lower().replace("\u00a0", " ")
            if k.startswith("nom"): idx["nom"] = j
            elif k.startswith("ressource"): idx["ress"] = j
            elif k.startswith("libell"): idx["lib"] = j
            elif k.startswith("commentaire"): idx["com"] = j
            elif k.startswith("type d"): idx["type"] = j
            elif k.startswith("date de déb") or k.startswith("date de deb"): idx["deb"] = j
            elif k.startswith("date de fin"): idx["fin"] = j
            elif k.startswith("charge"): idx["charge"] = j
            elif k.startswith("tâche") or k.startswith("tache"): idx["tache"] = j
            elif k.startswith("code imputation") or k.startswith("imputation"): idx["imput"] = j
        if "nom" in idx and "deb" in idx:
            return i, idx
    return 0, {"nom":0,"ress":1,"lib":2,"com":3,"type":4,"charge":10,"deb":11,"fin":12,"tache":14,"imput":15}

def main():
    path = find_xlsx()
    wb = load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    # certains exports déclarent une dimension fausse (ex: A1:O10) : on force le recalcul
    try:
        ws.reset_dimensions()
    except Exception:
        pass
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    hr, idx = find_cols(rows)
    section = ""
    out = []
    def g(r, key):
        j = idx.get(key)
        return r[j] if (j is not None and j < len(r)) else None
    for r in rows[hr + 1:]:
        nom = cell(g(r, "nom"))
        date = to_dt(g(r, "deb"))
        if date is None:
            other = cell(g(r,"lib")) or cell(g(r,"com")) or cell(g(r,"type")) or cell(g(r,"ress"))
            if nom and nom.lower() != "nom-prénom" and not other:
                section = nom
            continue
        if not nom or nom.lower() == "nom-prénom":
            continue
        fin = to_dt(g(r, "fin"))
        out.append({
            "nom": nom,
            "date": date.strftime("%Y-%m-%dT%H:%M"),
            "end": fin.strftime("%Y-%m-%dT%H:%M") if fin else "",
            "lib": cell(g(r,"lib")), "com": cell(g(r,"com")), "section": section,
            "ress": cell(g(r,"ress")), "tache": cell(g(r,"tache")), "charge": cell(g(r,"charge")),
            "imput": cell(g(r,"imput")),
        })
    data = {"generated": datetime.now().strftime("%d/%m/%Y %H:%M"), "count": len(out), "entries": out}
    with open("planning-data.json", "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, separators=(",", ":"))
    print(f"OK : {len(out)} affectations -> planning-data.json (source : {path})")

if __name__ == "__main__":
    main()
